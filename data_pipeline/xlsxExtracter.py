import pandas as pd
from openpyxl import load_workbook


file_path = '/content/PPS sheets (1).xlsx'
wb = load_workbook(file_path, data_only=True)
sheets = wb.sheetnames


def extract_hyperlink(cell):
    if cell.hyperlink:
        return cell.hyperlink.target
    return None


all_modules = {}
module_counter = 1
for sheet in sheets:
    ws = wb[sheet]
    df = pd.DataFrame(ws.values)

    df.columns = df.iloc[0]
    df = df[1:]

    if 'Final Topic name ' in df.columns and 'Video link ' in df.columns:
        module_name = f"Module {module_counter}"
        all_modules[module_name] = []
        for row in ws.iter_rows(min_row=2, max_col=len(df.columns), values_only=False):
            name_cell = row[df.columns.get_loc('Final Topic name ')]
            link_cell = row[df.columns.get_loc('Video link ')]
            name = name_cell.value
            link = extract_hyperlink(link_cell)

  
            if not any(cell.value for cell in row):
                if all_modules[module_name]: 
                    module_counter += 1
                    module_name = f"Module {module_counter}"
                    all_modules[module_name] = []
            else:
                if link:
                    all_modules[module_name].append((name, link))


for module, links in all_modules.items():
    print(f"\n{module}:")
    for name, link in links:
        print(f"  Name: {name}, Video link: {link}")
